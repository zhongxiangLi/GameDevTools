from openpyxl import load_workbook
import os
import sys
from openpyxl import Workbook
from . import Helper
from prettytable import PrettyTable


def ShowReadme():
        Helper.ShowReadme("数据类型限定:int string table")
        tmpTable=PrettyTable([u"成就ID",u"成就名称",u"奖励"])
        tmpTable.align[u"成就ID"]="1" #以这个字段左对齐
        tmpTable.padding_width=4 # 填充宽度

        tmpTable.add_row(["id","name","praise"])
        tmpTable.add_row(["int","string","table"])
        tmpTable.add_row([70101,u"提升等级","{{1000001,1}}"])
        tmpTable.add_row([70102,u"装备等级","{{1000001,1},{1000002,50}}"])
        print(tmpTable)

def run():
        ShowReadme()
        
        tmpExcelPath=input(u"输入Excel文件路径:")
        # tmpExcelPath="G:/achievement.xlsx"
        tmpExcelPath=tmpExcelPath.replace("\r","").replace("\n","")

        tmpVarNameChineseInRow=-1#注释在第4行
        tmpVarNameInRow=-1#字段名在第5行
        tmpVarTypedefInRow=-1#字段类型在第6行
        tmpDataBeginRow=-1#数据从第7行开始

        while tmpVarNameInRow==-1:
                tmpVarNameInRow=Helper.String_to_int(input(u"字段名 在第几行:"),-1)
        
        print(u"字段名 行号:"+str(tmpVarNameInRow)+"\n")

        while tmpVarTypedefInRow==-1:
                tmpVarTypedefInRow=Helper.String_to_int(input(u"字段类型 在第几行:"),-1)
        print(u"字段类型 行号:"+str(tmpVarTypedefInRow)+"\n")

        tmpVarNameChineseInRow=Helper.String_to_int(input(u"字段注释 在第几行,默认-1:"),-1)
        print(u"字段注释 行号:"+str(tmpVarNameChineseInRow)+"\n")

        while tmpDataBeginRow==-1:
                tmpDataBeginRow=Helper.String_to_int(input(u"正式数据 在第几行开始:"),-1)
        print(u"正式数据 行号:"+str(tmpDataBeginRow)+"\n")
        
        
        #整个Excel表格，导出到一个lua文件
        tmpLuaCodeStr=""

        #打开Excel表格
        tmpWorkBook=load_workbook(tmpExcelPath)
        for tmpSheetName in tmpWorkBook.sheetnames:
                #处理Excel文件中的一个数据表Sheet
                print("Convert "+tmpSheetName)

                #lua 数据表的开始
                tmpLuaData_CodeStr=tmpSheetName+"=\n{\n"

                tmpWorkBook_sheet=tmpWorkBook[tmpSheetName]
                tmpRows=tmpWorkBook_sheet.iter_rows()

                #字段注释 行
                tmpVarNameChineseRow=()

                #字段名 行
                tmpVarNameRow=()

                #字段类型 行
                tmpVarTypedefRow=()

                #遍历 对每一行数据进行处理
                tmpRowIndex=1
                for tmpRow in tmpRows:
                        if tmpRowIndex<tmpDataBeginRow:
                                # print("skip\n")
                                if tmpRowIndex==tmpVarNameChineseInRow:#字段注释 行
                                        tmpVarNameChineseRow=tmpRow
                                if tmpRowIndex==tmpVarNameInRow:#字段名 行
                                        tmpVarNameRow=tmpRow
                                if tmpRowIndex==tmpVarTypedefInRow:#字段类型 行
                                        tmpVarTypedefRow=tmpRow
                        else:
                                #一行一个table 数组形式,节省内存 类似于 {15,"激活1个先天灵宝"}
                                tmpRowStr="    {"

                                #对一行数据遍历，每一个Cell，然后拼成lua table
                                tmpAllCellNone=True
                                #如果这一行每个Cell都是None，那么说明到了表格的结尾，后续行 不管有没有数据，都不管了。
                                for tmpCellIndex in range(len(tmpRow)):
                                        tmpCell=tmpRow[tmpCellIndex]#当前Cell
                                        tmpCell_VarType=str(tmpVarTypedefRow[tmpCellIndex].value)#判断当前Cell数据类型

                                        if tmpCell_VarType=="string":
                                                tmpRowStr=tmpRowStr+'"'+str(tmpCell.value)+'",'
                                        elif tmpCell_VarType=="int":
                                                tmpRowStr=tmpRowStr+str(tmpCell.value)+','
                                        elif tmpCell_VarType=="float":
                                                tmpRowStr=tmpRowStr+str(tmpCell.value)+','
                                        elif tmpCell_VarType=="table":
                                                tmpRowStr=tmpRowStr+str(tmpCell.value)+','

                                        if str(tmpCell.value)!="None":
                                                tmpAllCellNone=False

                                
                                tmpRowStr=tmpRowStr[0:-1]
                                tmpRowStr=tmpRowStr+"},\n"
                                # print(tmpRowStr)

                                if tmpAllCellNone:#如果这一行每个Cell都是None，那么说明到了表格的结尾，后续行 不管有没有数据，都不管了。
                                        break
                                        
                                tmpLuaData_CodeStr=tmpLuaData_CodeStr+tmpRowStr

                        tmpRowIndex+=1

                #lua 数据表的结尾
                tmpLuaData_CodeStr=tmpLuaData_CodeStr+"}"
                # print(tmpLuaData_CodeStr)

        
                #lua 数据表字段 table，把字段保存为table 映射到数据表的 index
                #把 字段名 这一行，保存为字段 table luacode
                '''
                achievement_group_key=
                {
                        id=0,--成就组ID
                        name=1,--成就组名称
                }
                '''
                #lua 数据表的开始
                tmpLuaKey_CodeStr=tmpSheetName+"_key=\n{\n"
                for tmpCellIndex in range(len(tmpVarNameRow)):
                        tmpCell=tmpVarNameRow[tmpCellIndex]#当前Cell

                        tmpRowStr="    "+str(tmpCell.value)+"="+str(tmpCellIndex)+","

                        #判断是否有注释以及类型
                        if len(tmpVarTypedefRow)>0 or len(tmpVarNameChineseRow)>0:
                                tmpRowStr=tmpRowStr+"--"
                                if len(tmpVarTypedefRow)>0:
                                        tmpRowStr=tmpRowStr+"["+str(tmpVarTypedefRow[tmpCellIndex].value)+"] "
                                if len(tmpVarNameChineseRow)>0:
                                        tmpRowStr=tmpRowStr+str(tmpVarNameChineseRow[tmpCellIndex].value)
                        tmpRowStr=tmpRowStr+"\n"

                        tmpLuaKey_CodeStr=tmpLuaKey_CodeStr+tmpRowStr
                tmpLuaKey_CodeStr=tmpLuaKey_CodeStr+"}"

                tmpLuaCodeStr=tmpLuaCodeStr+tmpLuaKey_CodeStr+"\n\n"+tmpLuaData_CodeStr+"\n\n"

        #保存 luacode
        tmpSaveFilrPath=Helper.ChangePathExt(tmpExcelPath,".lua")
        Helper.WriteStrToTxtFile(tmpLuaCodeStr,tmpSaveFilrPath)

if __name__=="__main__":
    run()
